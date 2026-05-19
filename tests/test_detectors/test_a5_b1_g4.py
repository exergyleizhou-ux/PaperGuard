"""A5、B1、G4 检测器测试。"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from paperguard.core.types import Severity
from paperguard.detectors.a5_decimal_consistency import A5DecimalConsistencyDetector
from paperguard.detectors.b1_grim import B1GRIMDetector, GRIMInput
from paperguard.detectors.g4_metadata_forensics import (
    G4MetadataForensicsDetector,
    MetadataForensicsInput,
)


def test_a5_flags_repeated_fraction() -> None:
    df = pd.DataFrame({"col": [1.48, 2.48, 3.48, 1.48, 2.48] * 13 + [3.48]})
    detector = A5DecimalConsistencyDetector()
    result = detector.detect(df, seed=42)
    assert result.applicable
    assert any(f.severity >= Severity.SUSPICIOUS for f in result.findings)


def test_a5_passes_diverse_decimals(genuine_data: pd.DataFrame) -> None:
    detector = A5DecimalConsistencyDetector()
    result = detector.detect(genuine_data, seed=42)
    severe = [f for f in result.findings if f.severity >= Severity.SUSPICIOUS]
    assert len(severe) == 0


def test_b1_grim_catches_impossible_mean() -> None:
    """Likert 1-5 量表，N=10，均值 3.15 不可能 (3.1 或 3.2 可能)。"""
    detector = B1GRIMDetector()
    inputs = [GRIMInput(mean=3.15, n=10, decimal_places=2, label="Q1")]
    result = detector.detect(inputs, seed=42)
    assert result.applicable
    assert len(result.findings) == 1
    assert result.findings[0].severity >= Severity.CONCERN


def test_b1_grim_accepts_valid_mean() -> None:
    """均值 3.20 with N=10 完全有效。"""
    detector = B1GRIMDetector()
    inputs = [GRIMInput(mean=3.20, n=10, decimal_places=2, label="Q1")]
    result = detector.detect(inputs, seed=42)
    assert result.applicable
    assert len(result.findings) == 0


def test_g4_detects_time_mismatch(tmp_path: Path) -> None:
    """文件创建于运行时，但声称实验始于 2020-01。"""
    excel_path = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to get active worksheet")
    ws["A1"] = "Sample"
    ws["B1"] = "Value"
    for i in range(2, 20):
        ws[f"A{i}"] = f"S{i}"
        ws[f"B{i}"] = i * 0.1
    wb.save(excel_path)

    # G4 检查 1 的语义："文件创建时间早于声称的实验开始时间"
    # （= 数据在还没做实验时就存在）→ CRITICAL。
    # 把 claimed_experiment_start 推到未来，新文件就早于它，触发 CRITICAL。
    inputs = MetadataForensicsInput(
        file_path=excel_path,
        claimed_experiment_start=datetime(2099, 1, 1),
        claimed_experiment_end=datetime(2099, 6, 30),
    )
    detector = G4MetadataForensicsDetector()
    result = detector.detect(inputs, seed=42)
    assert result.applicable
    critical = [f for f in result.findings if f.severity == Severity.CRITICAL]
    assert len(critical) >= 1


def test_g4_detects_creator_mismatch(tmp_path: Path) -> None:
    """openpyxl 写出的文件 creator='openpyxl'，与声称作者不匹配 → CONCERN。"""
    excel_path = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to get active worksheet")
    ws["A1"] = "x"
    wb.save(excel_path)

    inputs = MetadataForensicsInput(
        file_path=excel_path,
        claimed_authors=["Alice Smith", "Bob Jones"],
    )
    result = G4MetadataForensicsDetector().detect(inputs, seed=42)
    assert result.applicable
    concerns = [
        f
        for f in result.findings
        if f.severity == Severity.CONCERN and "creator" in f.summary.lower()
    ]
    assert len(concerns) >= 1
